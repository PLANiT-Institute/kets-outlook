# -*- coding: utf-8 -*-
"""Reproduce the CCE reserve-price scenarios on one consistent baseline.

The model starts from P0/L0 (no reserve carve-out) and replaces the legacy
70 Mt bank assumption with the observed 2024 registry total. The reported
quantity is *required withholding*: the static supply adjustment needed to
support the conditional floor path. It is not an estimate of auction non-sales,
because the model does not contain an auction-demand curve.

Run: PYTHONPATH=. python3 scripts/run_escalator_floor.py
Output: outputs/runs/escalator_floor_cce_v2.0.json
"""

from copy import deepcopy
import json
import os
import unicodedata

import openpyxl

import sys
from pathlib import Path as _Path
sys.path.insert(0, str(_Path(__file__).resolve().parents[1]))  # 리포 루트 = kets/ 패키지

from kets.engine import KETSModel
from kets.excel_source import load_data_from_excel


MODEL_INPUTS_JSON = os.environ.get("CCE_MODEL_INPUTS_JSON")
if MODEL_INPUTS_JSON:
    with open(MODEL_INPUTS_JSON, encoding="utf-8") as input_file:
        DATA = json.load(input_file)
else:
    DATA = load_data_from_excel()
SID = "base"
YEARS = list(range(2026, 2041))
OUT_JSON = os.environ.get(
    "CCE_FLOOR_OUT", "outputs/runs/escalator_floor_cce_v2.0.json"
)


def observed_bank_2024_Mt():
    """Aggregate the 2024 carry-over and offset carry-over registry fields."""
    if MODEL_INPUTS_JSON:
        return float(DATA["banking_initial_bank_Mt"])
    wb = openpyxl.load_workbook("data/ets_data.xlsx", read_only=True, data_only=True)
    target = unicodedata.normalize("NFC", "배출권이월량")
    sheet_name = next(
        name for name in wb.sheetnames if unicodedata.normalize("NFC", name) == target
    )
    rows = wb[sheet_name].iter_rows(values_only=True)
    next(rows)
    total = 0.0
    for row in rows:
        if row[4] == 2024:
            total += float(row[5] or 0.0) + float(row[6] or 0.0)
    wb.close()
    return total / 1e6


OBSERVED_BANK_2024_MT = observed_bank_2024_Mt()


def geo_pkg(f0, g, cost_scale=1.0):
    floor_path = {y: f0 * (1 + g) ** (y - 2026) for y in YEARS}
    # 기술비용 배율은 엔진의 cost_multiplier 레버로 준다 — 웹·MCP가 쓰는 것과 같은 경로.
    data = ({**DATA, "model_params": {**DATA["model_params"], "cost_multiplier": cost_scale}}
            if cost_scale != 1.0 else DATA)
    model = KETSModel(data, "step")
    model.B0 = OBSERVED_BANK_2024_MT * 1e6
    model.corridor_floor_path = lambda *args, **kwargs: [floor_path[y] for y in YEARS]

    # P0 is the no-policy counterfactual used for the paper's equilibrium path.
    # Using package A here would pre-carve the 85.28 Mt reserve and silently change
    # the baseline against which withholding is measured.
    package = deepcopy(next(p for p in model.packages if p["package_id"] == "P0"))
    package["package_id"] = "CCE_floor_scenario"
    package["floor_id"] = "corridor"
    return model, model.solve_package(package)


def extract(model, path):
    auction_share = model.auction_ratio_path(SID, "A_gov")
    prices = [row["kau"] for row in path]
    activation = model.tech_activation_years(prices, SID, headline_only=True)
    annual_withholding, auction_volume, sold_volume = {}, {}, {}

    for row in path:
        year = row["year"]
        auction_volume[year] = auction_share[year] * model.cap_total[SID][year] / 1e6
        withholding = 0.0
        if row["floor"] > row["kau_prefloor"]:
            withholding = max(
                model.abatement_at_price(row["kau"], year, SID)
                - model.abatement_at_price(row["kau_prefloor"], year, SID),
                0.0,
            ) / 1e6
        annual_withholding[year] = round(withholding, 3)
        sold_volume[year] = max(auction_volume[year] - withholding, 0.0)

    revenue = sum(
        sold_volume[row["year"]] * 1e6 * row["kau"] / 1e12 for row in path
    )
    steel_token = unicodedata.normalize("NFC", "수소환원")
    ncc_token = unicodedata.normalize("NFC", "전기분해")
    steel = next(
        (tech for tech in activation if steel_token in unicodedata.normalize("NFC", tech)),
        None,
    )
    ncc = next(
        (
            tech
            for tech in activation
            if "e-cracker" in tech or ncc_token in unicodedata.normalize("NFC", tech)
        ),
        None,
    )
    return {
        "floor": {row["year"]: round(row["floor"]) for row in path},
        "conditional_price": {row["year"]: round(row["kau"]) for row in path},
        "prefloor_price": {row["year"]: round(row["kau_prefloor"]) for row in path},
        "steel_threshold_year": activation.get(steel),
        "ncc_threshold_year": activation.get(ncc),
        "min_headroom": round(min(row["headroom"] for row in path), 3),
        "defended_all": all(row["defended"] for row in path),
        "cum_required_withholding_Mt": round(sum(annual_withholding.values()), 1),
        "max_annual_required_withholding_Mt": round(max(annual_withholding.values()), 1),
        "annual_required_withholding_Mt": annual_withholding,
        "auction_volume_Mt": {y: round(v, 3) for y, v in auction_volume.items()},
        "sold_auction_volume_Mt": {y: round(v, 3) for y, v in sold_volume.items()},
        "conditional_cum_auction_revenue_trillion": round(revenue, 1),
    }


COMBOS = {
    "40k_7pct": (40000, 0.07),
    "45k_7pct": (45000, 0.07),
    "50k_7pct": (50000, 0.07),
    "45k_5.5pct": (45000, 0.055),
    "50k_5.5pct": (50000, 0.055),
}

out = {
    "meta": {
        "version": "CCE 2.0",
        "baseline_package": "P0/L0 (no reserve carve-out)",
        "initial_bank_Mt": round(OBSERVED_BANK_2024_MT, 6),
        "interpretation": (
            "Required withholding is the model-implied static supply adjustment needed to "
            "support the conditional price path. It is not an estimate of auction non-sales "
            "because the model does not contain an auction-demand curve."
        ),
    }
}
for name, (f0, growth) in COMBOS.items():
    model, scenario_path = geo_pkg(f0, growth)
    out[name] = extract(model, scenario_path)

sensitivity = []
for f0 in (40000, 45000, 50000):
    for cost_scale in (0.8, 1.0, 1.2):
        model, scenario_path = geo_pkg(f0, 0.07, cost_scale=cost_scale)
        result = extract(model, scenario_path)
        sensitivity.append(
            {
                "f0": f0,
                "cost_scale": cost_scale,
                "cum_required_withholding_Mt": result["cum_required_withholding_Mt"],
                "defended_all": result["defended_all"],
            }
        )
out["cost_sensitivity_pm20"] = {
    "grid": sensitivity,
    "cum_required_withholding_min_Mt": min(
        row["cum_required_withholding_Mt"] for row in sensitivity
    ),
    "cum_required_withholding_max_Mt": max(
        row["cum_required_withholding_Mt"] for row in sensitivity
    ),
    "defended_all_combos": all(row["defended_all"] for row in sensitivity),
}

os.makedirs(os.path.dirname(OUT_JSON) or ".", exist_ok=True)
with open(OUT_JSON, "w", encoding="utf-8") as output_file:
    json.dump(out, output_file, ensure_ascii=False, indent=2)

print(
    f"{'scenario':12} {'2040 floor':>11} {'steel':>6} {'NCC':>6} "
    f"{'withhold Mt':>12} {'annual max':>10} {'headroom':>9}"
)
for name, result in out.items():
    if "floor" not in result:
        continue
    print(
        f"{name:12} {result['floor'][2040]:>11,} "
        f"{str(result['steel_threshold_year']):>6} {str(result['ncc_threshold_year']):>6} "
        f"{result['cum_required_withholding_Mt']:>12} "
        f"{result['max_annual_required_withholding_Mt']:>10} "
        f"{result['min_headroom']:>9}"
    )

if __name__ == "__main__":
    assert abs(OBSERVED_BANK_2024_MT - 92.140327) < 1e-6
    assert out["45k_7pct"]["floor"][2040] == 116034
    assert out["45k_7pct"]["defended_all"]
