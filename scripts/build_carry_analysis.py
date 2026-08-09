# -*- coding: utf-8 -*-
"""Reproduce the CCE adjacent-vintage and banking analysis.

The primary statistic is an adjacent-vintage premium, not an annualized return:
KAU vintages differ by one compliance label, but contract timing, convenience
yield, transaction costs, and stale quotes prevent a literal yield interpretation.

Outputs
-------
outputs/runs/carry_analysis_cce_v2.0.json
outputs/supplementary/carry_pairs_cce.csv
outputs/supplementary/carry_summary_by_year_cce.csv
"""

import csv
import json
import os
import random
import statistics
import unicodedata

import numpy as np
import openpyxl


XLSX = "data/ets_data.xlsx"
OUT_JSON = "outputs/runs/carry_analysis_cce_v2.0.json"
OUT_PAIRS = "outputs/supplementary/carry_pairs_cce.csv"
PAIR_CSV_FIELDS = [
    "date", "calendar_year", "vintage_front", "vintage_next",
    "close_front_krw", "close_next_krw", "volume_front", "volume_next",
    "vintage_premium", "both_traded", "within_pm40pct",
]
OUT_YEAR = "outputs/supplementary/carry_summary_by_year_cce.csv"
BENCHMARK = 0.055
BOOTSTRAP_DRAWS = 5000
SEED = 20260801
# Statutory borrowing ceiling: Enforcement Decree Article 36(2) caps borrowing at
# 10% of the allowances an entity must surrender under Act Article 27(1).
# A legal parameter, not a model input, so it lives with BENCHMARK rather than in
# the master workbook.
BORROW_CAP_SHARE = 0.10


def sheet_by_name(workbook, name):
    target = unicodedata.normalize("NFC", name)
    actual = next(
        sheet
        for sheet in workbook.sheetnames
        if unicodedata.normalize("NFC", sheet) == target
    )
    return workbook[actual]


def mean(values):
    return statistics.fmean(values) if values else None


def percentile(sorted_values, probability):
    if not sorted_values:
        return None
    position = (len(sorted_values) - 1) * probability
    lower = int(position)
    upper = min(lower + 1, len(sorted_values) - 1)
    weight = position - lower
    return sorted_values[lower] * (1 - weight) + sorted_values[upper] * weight


def cluster_bootstrap_ci(rows, statistic):
    """Date-cluster bootstrap; all pairs from a sampled date move together."""
    by_date = {}
    for row in rows:
        by_date.setdefault(row["date"], []).append(row)
    dates = sorted(by_date)
    rng = random.Random(SEED)
    estimates = []
    for _ in range(BOOTSTRAP_DRAWS):
        sample = []
        for sampled_date in rng.choices(dates, k=len(dates)):
            sample.extend(by_date[sampled_date])
        estimates.append(statistic([row["vintage_premium"] for row in sample]))
    estimates.sort()
    return [percentile(estimates, 0.025), percentile(estimates, 0.975)]


def summarize(rows):
    premiums = [row["vintage_premium"] for row in rows]
    weights = [row["volume_front"] + row["volume_next"] for row in rows]
    weight_total = sum(weights)
    return {
        "n_pairs": len(rows),
        "n_dates": len({row["date"] for row in rows}),
        "median_pct": round(100 * statistics.median(premiums), 3),
        "mean_pct": round(100 * mean(premiums), 3),
        "sd_pct": round(100 * statistics.stdev(premiums), 3) if len(rows) > 1 else None,
        "p10_pct": round(100 * percentile(sorted(premiums), 0.10), 3),
        "p90_pct": round(100 * percentile(sorted(premiums), 0.90), 3),
        "share_below_5_5_pct": round(sum(p < BENCHMARK for p in premiums) / len(rows), 4),
        "volume_weighted_mean_pct": (
            round(100 * sum(p * w for p, w in zip(premiums, weights)) / weight_total, 3)
            if weight_total > 0
            else None
        ),
    }


workbook = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

# KAU daily closes and volumes.
market = sheet_by_name(workbook, "market").iter_rows(values_only=True)
next(market)
daily, kau_quotes = {}, 0
for row in market:
    date, instrument, close, volume = str(row[0])[:10], row[1], row[2], row[8]
    if (
        isinstance(instrument, str)
        and instrument.startswith("KAU")
        and len(instrument) == 5
        and close is not None
    ):
        kau_quotes += 1
        daily.setdefault(date, {})[int(instrument[3:])] = (
            float(close),
            float(volume or 0),
        )

pairs = []
for date in sorted(daily):
    vintages = daily[date]
    for vintage in sorted(vintages):
        if vintage + 1 not in vintages:
            continue
        price_front, volume_front = vintages[vintage]
        price_next, volume_next = vintages[vintage + 1]
        if price_front <= 0:
            continue
        premium = price_next / price_front - 1
        pairs.append(
            {
                "date": date,
                "calendar_year": int(date[:4]),
                "vintage_front": vintage,
                "vintage_next": vintage + 1,
                "close_front_krw": price_front,
                "close_next_krw": price_next,
                "volume_front": volume_front,
                "volume_next": volume_next,
                "vintage_premium": premium,
                "both_traded": int(volume_front > 0 and volume_next > 0),
                "within_pm40pct": int(-0.40 <= premium <= 0.40),
            }
        )

traded = [row for row in pairs if row["both_traded"]]
# Compliance-distance test.
#
# Months to the front vintage's surrender deadline (31 August of the following
# year) is the one confound the pooled premium statistics cannot separate on their
# own. Conditioning on it raises the pooled mean near the deadline and lowers it
# away from the deadline, which looks like an identified compliance effect. It is
# not: with calendar-year fixed effects and date-clustered standard errors the
# proximity coefficient is not distinguishable from zero. The block is retained so
# that the negative result is reproducible and citable as a limitation.
def months_to_deadline(row):
    deadline_year, deadline_month = 2000 + row["vintage_front"] + 1, 8
    year, month = int(row["date"][:4]), int(row["date"][5:7])
    return (deadline_year - year) * 12 + (deadline_month - month)


for row in pairs:
    row["months_to_deadline"] = months_to_deadline(row)

traded_years = sorted({row["calendar_year"] for row in traded})


def _proximity_regression(rows, years):
    """p = year FE + b*near, date-clustered SE. Returns (b, se, t, n, n_clusters)."""
    design, outcome, clusters = [], [], []
    for row in rows:
        dummies = [1.0 if row["calendar_year"] == y else 0.0 for y in years]
        design.append(dummies + [1.0 if row["months_to_deadline"] <= 1 else 0.0])
        outcome.append(100 * row["vintage_premium"])
        clusters.append(row["date"])
    X, y = np.array(design), np.array(outcome)
    n, k = X.shape
    xtx_inv = np.linalg.pinv(X.T @ X)
    beta = xtx_inv @ X.T @ y
    resid = y - X @ beta
    meat = np.zeros((k, k))
    for cluster in set(clusters):
        idx = [i for i, c in enumerate(clusters) if c == cluster]
        score = X[idx].T @ resid[idx]
        meat += np.outer(score, score)
    groups = len(set(clusters))
    scale = (groups / (groups - 1)) * ((n - 1) / (n - k))
    se = float(np.sqrt((xtx_inv @ meat @ xtx_inv * scale)[-1, -1]))
    return float(beta[-1]), se, n, groups


prox_beta, prox_se, prox_n, prox_groups = _proximity_regression(traded, traded_years)

horizon_rows = {}
for row in traded:
    horizon_rows.setdefault(row["months_to_deadline"], []).append(row)

trimmed = [row for row in pairs if row["within_pm40pct"]]

# Forward-horizon trading activity.
#
# h = vintage year - calendar year. h <= 0 covers vintages whose compliance
# obligation is already fixed; h >= 1 is a claim on a future vintage. Hotelling
# transmission presupposes that the h >= 1 claim is tradable, so its realised
# volume measures whether the intertemporal arbitrage channel exists at all.
# This is a volume statistic and is therefore not exposed to the stale-quote,
# convenience-yield, or transaction-cost readings that the price premium is.
horizon = {}
forward_traded_days = []
identical_close_days = 0
forward_quote_days = 0
for date, vintages in sorted(daily.items()):
    year = int(date[:4])
    forward_closes = []
    for vintage, (close, volume) in vintages.items():
        h = 2000 + vintage - year
        bucket = horizon.setdefault(
            h, {"quote_days": 0, "traded_days": 0, "volume_tco2": 0.0}
        )
        bucket["quote_days"] += 1
        bucket["traded_days"] += int(volume > 0)
        bucket["volume_tco2"] += volume
        if h >= 1:
            forward_closes.append(close)
            if volume > 0:
                forward_traded_days.append(
                    {"date": date, "vintage": 2000 + vintage, "horizon": h,
                     "volume_tco2": volume}
                )
    if len(forward_closes) > 1:
        forward_quote_days += 1
        identical_close_days += int(len(set(forward_closes)) == 1)

forward = {h: value for h, value in horizon.items() if h >= 1}
spot = {h: value for h, value in horizon.items() if h <= 0}


def _totals(buckets):
    return {
        "quote_days": sum(v["quote_days"] for v in buckets.values()),
        "traded_days": sum(v["traded_days"] for v in buckets.values()),
        "volume_tco2": sum(v["volume_tco2"] for v in buckets.values()),
    }


forward_total = _totals(forward)
spot_total = _totals(spot)
all_volume = forward_total["volume_tco2"] + spot_total["volume_tco2"]

# Degenerate-observation accounting for the full quote-pair sample. A pair whose
# two legs print the same reference close contributes an exactly zero premium
# that carries no market information; the both-traded filter removes them.
# Pair composition by trading status. Pairs in which neither leg traded compare two
# reference prints: they are exactly zero when the prints coincide and drift apart
# otherwise. Neither branch carries market information, so the all-quote statistics
# are reported for completeness rather than as an economic contrast.
by_status = {"both_traded": [], "one_traded": [], "neither_traded": []}
for row in pairs:
    front, next_ = row["volume_front"] > 0, row["volume_next"] > 0
    key = ("both_traded" if front and next_
           else "neither_traded" if not front and not next_
           else "one_traded")
    by_status[key].append(row)


def status_block(rows):
    zeros = [row for row in rows if row["vintage_premium"] == 0.0]
    nonzero = sorted(row["vintage_premium"] for row in rows
                     if row["vintage_premium"] != 0.0)
    return {
        "n_pairs": len(rows),
        "n_exactly_zero": len(zeros),
        "share_exactly_zero": round(len(zeros) / len(rows), 4),
        "n_nonzero": len(nonzero),
        "nonzero_median_pct": round(100 * statistics.median(nonzero), 3) if nonzero else None,
        "nonzero_sd_pct": (round(100 * statistics.stdev(nonzero), 3)
                           if len(nonzero) > 1 else None),
        "nonzero_p10_pct": round(100 * percentile(nonzero, 0.10), 3) if nonzero else None,
        "nonzero_p90_pct": round(100 * percentile(nonzero, 0.90), 3) if nonzero else None,
    }


exact_zero = sum(1 for row in pairs if row["vintage_premium"] == 0.0)
forward_leg_pairs = sum(
    1 for row in pairs
    if row["vintage_front"] + 2000 > row["calendar_year"]
    or row["vintage_next"] + 2000 > row["calendar_year"]
)

# Annual descriptive results for the liquidity-filtered sample.
year_rows = []
for year in sorted({row["calendar_year"] for row in traded}):
    rows = [row for row in traded if row["calendar_year"] == year]
    year_rows.append({"calendar_year": year, **summarize(rows)})


def year_sum(sheet_name, value_columns):
    totals = {}
    rows = sheet_by_name(workbook, sheet_name).iter_rows(values_only=True)
    next(rows)
    for row in rows:
        if row[4] is None:
            continue
        year = int(row[4])
        totals[year] = totals.get(year, 0.0) + sum(
            float(row[column] or 0.0) for column in value_columns
        )
    return totals


# Auction outcomes.
#
# The paper models "required withholding" and is careful not to call it predicted
# non-sales. The auction record shows why that caution is right and where it can be
# strengthened: Korea already leaves a large share of offered volume unsold, but
# almost all of it comes from insufficient bids rather than from a binding reserve.
# The two channels are separated here so the distinction is reproducible.
auction_rows = []
_auction = sheet_by_name(workbook, "auction").iter_rows(values_only=True)
next(_auction)
for row in _auction:
    if row[0] is None:
        continue
    auction_rows.append({
        "date": str(row[0])[:10],
        "instrument": row[1],
        "offered_tco2": float(row[2] or 0),
        "bid_tco2": float(row[3] or 0),
        "bidders": int(row[5] or 0),
        "winners": int(row[6] or 0),
        "lowest_bid_krw": float(row[8] or 0),
        "sold_tco2": float(row[9] or 0),
        "clearing_krw": float(row[10] or 0),
    })

_offered = sum(a["offered_tco2"] for a in auction_rows)
_sold = sum(a["sold_tco2"] for a in auction_rows)
_undersold = [a for a in auction_rows if a["sold_tco2"] < a["offered_tco2"]]
_demand_short = [a for a in _undersold if a["bid_tco2"] < a["offered_tco2"]]
_covered_unsold = [a for a in _undersold if a["bid_tco2"] >= a["offered_tco2"]]
_cover = sorted(a["bid_tco2"] / a["offered_tco2"]
                for a in auction_rows if a["offered_tco2"] > 0)

bank = year_sum("배출권이월량", [5, 6])
verified = year_sum("인증배출량", [5])
borrow = year_sum("배출권차입량", [5])


def firm_years(sheet_name, firm_column=3, year_column=4):
    """(업체, 연도) 관측 수와 고유 업체 수 — 등록부 이용 폭의 표본크기."""
    observations, firms = 0, set()
    rows = sheet_by_name(workbook, sheet_name).iter_rows(values_only=True)
    next(rows)
    for row in rows:
        if row[year_column] is None:
            continue
        observations += 1
        firms.add(row[firm_column])
    return observations, len(firms)


borrow_obs, borrow_firms = firm_years("배출권차입량")
bank_obs, bank_firms = firm_years("배출권이월량")
banking_ratio = {
    year: bank[year] / verified[year]
    for year in sorted(bank)
    if year in verified and verified[year] > 0
}
workbook.close()

all_summary = summarize(pairs)
traded_summary = summarize(traded)
mean_ci = cluster_bootstrap_ci(traded, statistics.fmean)
median_ci = cluster_bootstrap_ci(traded, statistics.median)
traded_summary["date_cluster_bootstrap_mean_ci95_pct"] = [
    round(100 * value, 3) for value in mean_ci
]
traded_summary["date_cluster_bootstrap_median_ci95_pct"] = [
    round(100 * value, 3) for value in median_ci
]

stats = {
    "meta": {
        "version": "CCE 2.0",
        "benchmark_pct": 5.5,
        "bootstrap_draws": BOOTSTRAP_DRAWS,
        "bootstrap_seed": SEED,
        "interpretation": (
            "The adjacent-vintage premium is descriptive. It is not a pure annualized "
            "financing return because convenience yield, transaction costs, compliance "
            "timing, risk, and stale closes are not separately identified."
        ),
    },
    "kau_vintage_day_quotes": kau_quotes,
    "forward_horizon": {
        "definition": (
            "h = vintage year minus calendar year. h <= 0 is a vintage whose "
            "compliance obligation is already fixed; h >= 1 is a claim on a "
            "future vintage, which is what intertemporal arbitrage requires."
        ),
        "by_horizon": {
            str(h): {
                "quote_days": value["quote_days"],
                "traded_days": value["traded_days"],
                "traded_day_share": round(
                    value["traded_days"] / value["quote_days"], 6
                ),
                "volume_tco2": round(value["volume_tco2"]),
            }
            for h, value in sorted(horizon.items())
        },
        "forward_totals": {
            "quote_days": forward_total["quote_days"],
            "traded_days": forward_total["traded_days"],
            "volume_tco2": round(forward_total["volume_tco2"]),
            "share_of_all_kau_volume": round(
                forward_total["volume_tco2"] / all_volume, 8
            ),
        },
        "forward_traded_days": forward_traded_days,
        "forward_quote_days_with_multiple_vintages": forward_quote_days,
        "forward_days_with_identical_close_across_vintages": identical_close_days,
    },
    "degenerate_pairs": {
        "n_pairs_total": len(pairs),
        "n_exactly_zero_premium": exact_zero,
        "share_exactly_zero_premium": round(exact_zero / len(pairs), 6),
        "n_pairs_with_forward_leg": forward_leg_pairs,
        "n_both_traded": len(traded),
    },
    "intertemporal_channel": {
        "note": (
            "Banking and borrowing move allowances across vintages inside a "
            "firm's own registry account. They change quantities but generate "
            "no cross-vintage price. Hotelling equalisation requires "
            "price-forming market transactions, which the forward_horizon "
            "block shows are absent."
        ),
        "banking_firm_year_obs": bank_obs,
        "banking_distinct_firms": bank_firms,
        "borrowing_firm_year_obs": borrow_obs,
        "borrowing_distinct_firms": borrow_firms,
        "borrowing_total_tco2": round(sum(borrow.values())),
        "by_year": {
            str(year): {
                "banked_tco2": round(bank.get(year, 0.0)),
                "borrowed_tco2": round(borrow.get(year, 0.0)),
                "bank_to_borrow_ratio": (
                    round(bank[year] / borrow[year], 1)
                    if borrow.get(year) else None
                ),
            }
            for year in sorted(set(bank) | set(borrow))
        },
        "borrowing_cap": {
            "share_of_surrender_obligation": BORROW_CAP_SHARE,
            "legal_basis": "Enforcement Decree Art. 36(2); Act Art. 27(1)",
            "by_year": {
                str(year): {
                    "ceiling_tco2": round(BORROW_CAP_SHARE * verified[year]),
                    "borrowed_tco2": round(borrow.get(year, 0.0)),
                    "utilisation": round(
                        borrow.get(year, 0.0) / (BORROW_CAP_SHARE * verified[year]), 6
                    ),
                }
                for year in sorted(verified)
                if verified[year] > 0
            },
        },
        "borrowing_to_forward_market_volume_ratio": round(
            sum(borrow.values()) / forward_total["volume_tco2"], 1
        ) if forward_total["volume_tco2"] else None,
    },
    "pair_composition_by_trading_status": {
        key: status_block(rows) for key, rows in by_status.items()
    },
    "compliance_distance_test": {
        "note": (
            "Negative result, retained for reproducibility. Pooled statistics split "
            "by months to the front vintage's surrender deadline suggest a "
            "proximity effect, but it does not survive calendar-year fixed effects."
        ),
        "deadline_rule": "31 August of the year following the front vintage",
        "by_months_to_deadline": {
            str(months): {
                "n_pairs": len(rows_m),
                "median_pct": round(100 * statistics.median(
                    [r["vintage_premium"] for r in rows_m]), 3),
                "mean_pct": round(100 * mean(
                    [r["vintage_premium"] for r in rows_m]), 3),
                "calendar_years": sorted({r["calendar_year"] for r in rows_m}),
            }
            for months, rows_m in sorted(horizon_rows.items())
        },
        "year_fe_proximity_regression": {
            "spec": "premium_pct ~ calendar-year FE + 1[months_to_deadline <= 1]",
            "coefficient_pp": round(prox_beta, 3),
            "cluster_robust_se_pp": round(prox_se, 3),
            "t_stat": round(prox_beta / prox_se, 3),
            "ci95_pp": [round(prox_beta - 1.96 * prox_se, 3),
                        round(prox_beta + 1.96 * prox_se, 3)],
            "n_pairs": prox_n,
            "n_date_clusters": prox_groups,
            "identifying_years": [
                y for y in traded_years
                if any(r["calendar_year"] == y and r["months_to_deadline"] <= 1 for r in traded)
                and any(r["calendar_year"] == y and r["months_to_deadline"] >= 2 for r in traded)
            ],
            "verdict": "not distinguishable from zero",
        },
    },
    "auction_outcomes": {
        "note": (
            "Undersubscription is decomposed into insufficient bids and bids that "
            "were sufficient but not fully allocated. Only the second channel is "
            "comparable to a reserve-price-driven non-sale, and it is small."
        ),
        "n_auctions": len(auction_rows),
        "first_date": auction_rows[0]["date"],
        "last_date": auction_rows[-1]["date"],
        "n_fully_sold": len(auction_rows) - len(_undersold),
        "n_undersold": len(_undersold),
        "n_zero_sold": sum(1 for a in auction_rows if a["sold_tco2"] == 0),
        "offered_tco2": round(_offered),
        "sold_tco2": round(_sold),
        "unsold_tco2": round(_offered - _sold),
        "unsold_share": round((_offered - _sold) / _offered, 6),
        "bid_to_cover": {
            "median": round(percentile(_cover, 0.50), 4),
            "mean": round(mean(_cover), 4),
            "min": round(_cover[0], 4),
            "max": round(_cover[-1], 4),
            "n": len(_cover),
        },
        "undersold_decomposition": {
            "insufficient_bids": {
                "n_auctions": len(_demand_short),
                "unsold_tco2": round(sum(a["offered_tco2"] - a["sold_tco2"]
                                         for a in _demand_short)),
            },
            "bids_sufficient_but_unsold": {
                "n_auctions": len(_covered_unsold),
                "unsold_tco2": round(sum(a["offered_tco2"] - a["sold_tco2"]
                                         for a in _covered_unsold)),
                "cases": [
                    {k: (round(v) if isinstance(v, float) else v)
                     for k, v in a.items()}
                    for a in _covered_unsold
                ],
            },
        },
    },
    "all_quote_pairs": all_summary,
    "both_traded_pairs": traded_summary,
    "legacy_pm40pct_display_sample_n": len(trimmed),
    "annual_both_traded": year_rows,
    "years_with_median_below_5_5_pct": sum(
        row["median_pct"] < 5.5 for row in year_rows
    ),
    "years_observed": len(year_rows),
    "banking_ratio": {year: round(value, 6) for year, value in banking_ratio.items()},
    "banking_ratio_2015_21_mean": round(
        mean([banking_ratio[year] for year in range(2015, 2022)]), 6
    ),
    "bank_2024_Mt": round(bank[2024] / 1e6, 6),
    "verified_emissions_2024_Mt": round(verified[2024] / 1e6, 6),
}

assert stats["kau_vintage_day_quotes"] == 8513
assert stats["all_quote_pairs"]["n_pairs"] == 5693
assert stats["both_traded_pairs"]["n_pairs"] == 253
assert stats["legacy_pm40pct_display_sample_n"] == 5444
assert abs(stats["bank_2024_Mt"] - 92.140327) < 1e-6
assert round(stats["banking_ratio"][2024], 3) == 0.168

os.makedirs(os.path.dirname(OUT_PAIRS), exist_ok=True)
with open(OUT_PAIRS, "w", newline="", encoding="utf-8") as file:
    # 공개 열 고정. months_to_deadline 같은 파생열은 JSON에만 둔다 —
    # 이 CSV는 SHA-256이 공표된 복제 패키지 산출물이라 열 추가가 검증을 깬다.
    writer = csv.DictWriter(file, fieldnames=PAIR_CSV_FIELDS,
                            extrasaction="ignore")
    writer.writeheader()
    writer.writerows(pairs)

with open(OUT_YEAR, "w", newline="", encoding="utf-8") as file:
    writer = csv.DictWriter(file, fieldnames=list(year_rows[0]))
    writer.writeheader()
    writer.writerows(year_rows)

with open(OUT_JSON, "w", encoding="utf-8") as file:
    json.dump(stats, file, ensure_ascii=False, indent=2)

print(json.dumps({key: value for key, value in stats.items() if key != "banking_ratio"}, indent=2))
print("saved:", OUT_JSON, OUT_PAIRS, OUT_YEAR)
